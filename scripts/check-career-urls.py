"""Check status of career site URLs in career_sites_with_ats.xlsx."""
import openpyxl
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib3.exceptions import InsecureRequestWarning
import warnings
import time

warnings.simplefilter("ignore", InsecureRequestWarning)

INPUT_FILE = "scripts/output/career_sites_with_ats.xlsx"
OUTPUT_FILE = "scripts/output/career_sites_with_status.xlsx"
TIMEOUT = 10
MAX_WORKERS = 30
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}


def check_url(idx: int, company: str, url: str):
    if not url or not isinstance(url, str):
        return idx, company, url, "missing", "", ""
    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True, verify=False, stream=True)
        r.close()
        status = r.status_code
        final = r.url
        if 200 <= status < 300:
            return idx, company, url, "ok", str(status), final
        if 300 <= status < 400:
            return idx, company, url, "redirect", str(status), final
        if status in (401, 403):
            return idx, company, url, "blocked", str(status), final
        if status in (404, 410):
            return idx, company, url, "dead", str(status), final
        return idx, company, url, "error", str(status), final
    except requests.exceptions.SSLError:
        return idx, company, url, "ssl_error", "", ""
    except requests.exceptions.ConnectionError:
        return idx, company, url, "dns_or_conn", "", ""
    except requests.exceptions.Timeout:
        return idx, company, url, "timeout", "", ""
    except Exception as e:
        return idx, company, url, "error", "", str(e)[:80]


def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"Checking {len(rows)} URLs with {MAX_WORKERS} workers...")

    results = [None] * len(rows)
    started = time.time()
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(check_url, i, r[0], r[1]): i for i, r in enumerate(rows)}
        done = 0
        for fut in as_completed(futures):
            idx, company, url, status, code, final = fut.result()
            results[idx] = (company, url, status, code, final)
            done += 1
            if done % 50 == 0:
                elapsed = time.time() - started
                print(f"  {done}/{len(rows)} done ({elapsed:.1f}s)")

    out = openpyxl.Workbook()
    sw = out.active
    sw.title = "Career Sites Status"
    sw.append(["Company", "Direct Job Board URL", "Status", "HTTP Code", "Final URL"])
    counts = {}
    for r in results:
        sw.append(list(r))
        counts[r[2]] = counts.get(r[2], 0) + 1
    out.save(OUTPUT_FILE)

    print()
    print(f"Total: {len(rows)}")
    for k in sorted(counts, key=lambda x: -counts[x]):
        print(f"  {k}: {counts[k]}")
    print(f"\nWrote: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
